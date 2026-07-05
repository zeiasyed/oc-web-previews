"""Parse OpenClinica CDASH Excel forms into schema dicts."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

try:
    import xlrd
except ImportError:
    xlrd = None

SKIP_TYPES = {"calculate", "settings", "note"}


def _choice_list(ftype: str, choices: dict[str, list[dict]]) -> list[dict]:
    m = re.match(r"select_(?:one|multiple)\s+(\S+)", ftype)
    if not m:
        return []
    return choices.get(m.group(1), [])


def parse_excel(path: Path) -> dict[str, Any]:
    if path.suffix.lower() == ".xls":
        return _parse_xls(path)
    return _parse_xlsx(path)


def _parse_xlsx(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    title, form_id = "Case Report Form", path.stem.split()[0]
    if "settings" in wb.sheetnames:
        ws0 = wb["settings"]
        hdr = [str(c).strip() if c else "" for c in next(ws0.iter_rows(max_row=1, values_only=True))]
        row1 = next(ws0.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if row1:
            meta = dict(zip(hdr, row1))
            title = str(meta.get("form_title") or title)
            form_id = str(meta.get("form_id") or form_id)

    choices: dict[str, list[dict]] = {}
    if "choices" in wb.sheetnames:
        ws = wb["choices"]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if not headers:
                headers = [str(c).strip() if c else "" for c in row]
                continue
            if not row or not row[0]:
                continue
            d = dict(zip(headers, row))
            lst = str(d.get("list_name") or "")
            if not lst:
                continue
            choices.setdefault(lst, []).append(
                {"code": str(d.get("name") or ""), "label": str(d.get("label") or d.get("name") or "")}
            )

    fields: list[dict] = []
    ws = wb["survey"]
    headers = None
    for row in ws.iter_rows(values_only=True):
        if not headers:
            headers = [str(c).strip() if c else "" for c in row]
            continue
        d = dict(zip(headers, row))
        ftype = str(d.get("type") or "").strip()
        name = str(d.get("name") or "").strip()
        if ftype in ("begin group", "end group", "begin repeat", "end repeat") or not ftype:
            continue
        if ftype in SKIP_TYPES:
            continue
        label = str(d.get("label") or d.get("Notes") or name)
        required = str(d.get("required") or "").lower() in ("yes", "true", "1")
        fields.append(
            {
                "name": name,
                "label": label,
                "type": ftype,
                "required": required,
                "choices": _choice_list(ftype, choices),
            }
        )
    return {"form_id": form_id, "title": title, "source": str(path), "fields": fields}


def _parse_xls(path: Path) -> dict[str, Any]:
    if xlrd is None:
        raise RuntimeError("xlrd required for .xls files")
    wb = xlrd.open_workbook(str(path))
    title, form_id = "Case Report Form", path.stem.split()[0]

    choices: dict[str, list[dict]] = {}
    if "choices" in wb.sheet_names():
        ws = wb.sheet_by_name("choices")
        headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
        for r in range(1, ws.nrows):
            row = [ws.cell_value(r, c) for c in range(ws.ncols)]
            d = dict(zip(headers, row))
            lst = str(d.get("list_name") or "")
            if not lst:
                continue
            choices.setdefault(lst, []).append(
                {"code": str(d.get("name") or ""), "label": str(d.get("label") or d.get("name") or "")}
            )

    fields: list[dict] = []
    ws = wb.sheet_by_name("survey")
    headers = [str(ws.cell_value(0, c)).strip() for c in range(ws.ncols)]
    for r in range(1, ws.nrows):
        row = [ws.cell_value(r, c) for c in range(ws.ncols)]
        d = dict(zip(headers, row))
        ftype = str(d.get("type") or "").strip()
        name = str(d.get("name") or "").strip()
        if ftype in ("begin group", "end group", "begin repeat", "end repeat") or not ftype:
            continue
        if ftype in SKIP_TYPES:
            continue
        label = str(d.get("label") or d.get("Notes") or name)
        required = str(d.get("required") or "").lower() in ("yes", "true", "1")
        fields.append(
            {
                "name": name,
                "label": label,
                "type": ftype,
                "required": required,
                "choices": _choice_list(ftype, choices),
            }
        )
    return {"form_id": form_id, "title": title, "source": str(path), "fields": fields}


def load_schema(form_code: str, schema_dir: Path) -> dict[str, Any] | None:
    path = schema_dir / f"{form_code}.json"
    if path.exists():
        return json.loads(path.read_text(encoding="utf-8"))
    return None


def save_schema(schema: dict[str, Any], form_code: str, schema_dir: Path) -> Path:
    schema_dir.mkdir(parents=True, exist_ok=True)
    out = schema_dir / f"{form_code}.json"
    out.write_text(json.dumps(schema, indent=2), encoding="utf-8")
    return out
